# redirect-engine.py
from openpyxl import load_workbook
from flask import Flask, request, redirect

app = Flask(__name__)

@app.route('/', methods=['GET'])
def handle_redirect():
    host = request.args.get('host')
    uri = request.args.get('uri')
    print(host + uri)
    # Look up the redirect URL
    redirect_url = find_redirect_url(host, uri)
    return redirect(redirect_url, 302)

def find_redirect_url(host, uri):
    workbook = load_workbook('RedirectLibrary.xlsx')
    if 'Redirect Library' not in workbook.sheetnames:
        return "<h1>\"Redirect Library\" sheet not found."
    worksheet = workbook['Redirect Library']
    redirectUrlIndex = binary_search(worksheet, host + uri)
    if redirectUrlIndex == -1:
        return "https://online.citi.com/US/ag/pageNotFound"
    return worksheet.cell(row=redirectUrlIndex, column=2).value

def binary_search(worksheet, requestUrl):
    left = 2
    right = find_index_of_last_row(worksheet)

    if right == 0:
        return -1

    while left <= right:
        mid = (left + right) // 2
        comparisonUrl = worksheet.cell(row=mid, column=1).value
        if comparisonUrl == requestUrl:
            return mid
        elif comparisonUrl < requestUrl:
            left = mid + 1
        else:
            right = mid - 1
    return -1

def find_index_of_last_row(worksheet):
    for row_idx in range(worksheet.max_row, 0, -1):
        row_is_empty = all(cell.value is None for cell in worksheet[row_idx])
        if not row_is_empty:
            # Found the real last non-empty row
            return row_idx
    # Sheet is empty
    return 0

if __name__ == "__main__":
    app.run(host="127.0.0.1", port=5000)
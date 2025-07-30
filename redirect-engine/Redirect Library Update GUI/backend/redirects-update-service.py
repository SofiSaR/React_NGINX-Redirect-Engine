# Backend for Update Redirect Library
# This Flask app provides REST API endpoints to view, add, edit, and delete Request URL - Redirect URL pairs in RedirectLibrary.xlsx.
# Place this file in the 'backend' folder. Requirements: Flask, openpyxl.

from flask import Flask, request, jsonify
from openpyxl import load_workbook, Workbook
from flask_cors import CORS
import os
from passlib.hash import bcrypt

app = Flask(__name__)
CORS(app)

EXCEL_FILE = os.path.join(os.path.dirname(__file__), '..', '..', 'RedirectLibrary.xlsx')

# Utility to load or create the Excel file
def get_redirects_sheet():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Redirect Library'
        ws.append(["Request URL", "Redirect URL"])
        wb.save(EXCEL_FILE)
    wb = load_workbook(EXCEL_FILE)
    if 'Redirect Library' not in wb.sheetnames:
        ws = wb.create_sheet('Redirect Library')
        ws.append(["Request URL", "Redirect URL"])
        wb.save(EXCEL_FILE)
    return wb, wb['Redirect Library']

def get_credentials_sheet():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Login Credentials'
        ws.append(["Username", "Password"])
        wb.save(EXCEL_FILE)
    wb = load_workbook(EXCEL_FILE)
    if 'Login Credentials' not in wb.sheetnames:
        ws = wb.create_sheet('Login Credentials')
        ws.append(["Username", "Password"])
        wb.save(EXCEL_FILE)
    return wb, wb['Login Credentials']


# Get all entries
@app.route('/api/redirects', methods=['GET'])
def get_redirects():
    wb, ws = get_redirects_sheet()
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            data.append({"request_url": row[0], "redirect_url": row[1]})
    return jsonify(data)
    
@app.route('/api/login', methods=['POST'])
def login():
    jsonParams = request.get_json()
    entered_username = jsonParams.get('username', '').strip()
    entered_password = jsonParams.get('password', '').strip()

    if not entered_username or not entered_password:
        return jsonify({
            "success": False,
            "message": "Please fill in required fields.",
        })

    wb, ws = get_credentials_sheet()

    username = ws.cell(row=2, column=1).value
    password = ws.cell(row=2, column=2).value

    if entered_username == username and bcrypt.verify(entered_password, password):
        return jsonify({
            "success": True,
            "message": "Login successful.",
        })
    else:
        return jsonify({
            "success": False,
            "message": "Invalid username or password.",
        })

# Handles any requests to alter the redirect library
@app.route('/api/redirects', methods=['POST'])
def add_redirect():
    jsonParams = request.get_json()
    action = jsonParams.get('action', '').strip()

    if not action:
        return jsonify({
            "success": False,
            "message": "Action not received by backend.",
            "index": -1
        })


    if action == 'update':
        wb, ws = get_redirects_sheet()
        request_url = jsonParams.get('request_url', '').strip()
        redirect_url = jsonParams.get('redirect_url', '').strip()
        if not request_url or not redirect_url:
            return jsonify({
                "success": False,
                "message": "URLs not received by backend.",
                "index": -1
            })
        index = jsonParams.get('index', None)
        if index is not None:
            index = int(index)
        else:
            init_request_url = jsonParams.get('init_request_url', '').strip()
            if not init_request_url:
                return jsonify({
                    "success": False,
                    "message": "Neither the index nor the initial request url of the entry to be edited were received by the backend.",
                    "index": -1
                })
            index = binary_search(ws, init_request_url)
            
        if ws.cell(row=index, column=1).value is not request_url:
            ws.delete_rows(index)
            [success, index] = binary_insert(ws, request_url)
            if not success:
                return jsonify({
                    "success": False,
                    "message": "Failed to update entry. You tried to update the request URL for this entry to one that already exists at index " + str(index - 1) + ".",
                    "index": index - 1
                })
            ws.insert_rows(index)
        ws.cell(row=index, column=1).value = request_url
        ws.cell(row=index, column=2).value = redirect_url
        wb.save(EXCEL_FILE)

        return jsonify({
            "success": True,
            "message": "Entry successfully updated.",
            "index": index - 1
        })
    

    if action == 'add':
        request_url = jsonParams.get('request_url', '').strip()
        redirect_url = jsonParams.get('redirect_url', '').strip()
        if not request_url or not redirect_url:
            return jsonify({
                "success": False,
                "message": "URLs not received by backend.",
                "index": -1
            })
        
        wb, ws = get_redirects_sheet()
        [success, index] = binary_insert(ws, request_url)
        if not success:
            return jsonify({
                "success": False,
                "message": "Request URL already exists at index " + str(index - 1),
                "index": index - 1
            })
        
        ws.insert_rows(index)
        ws.cell(row=index, column=1).value = request_url
        ws.cell(row=index, column=2).value = redirect_url
        wb.save(EXCEL_FILE)

        return jsonify({
            "success": True,
            "message": "Entry successfully added.",
            "index": index - 1
        })
    

    if action == 'delete':
        wb, ws = get_redirects_sheet()
        index = jsonParams.get('index', None)
        if index is not None:
            index = int(index)
        else:
            request_url = jsonParams.get('request_url', '').strip()
            if request_url:
                index = binary_search(ws, request_url)
                if index is None:
                    return jsonify({
                        "success": False,
                        "message": "Index of delete entry could not be retrieved.",
                        "index": -1
                    })
            else:
                return jsonify({
                    "success": False,
                    "message": "Neither the index nor the request url of the entry to be deleted were received by the backend.",
                    "index": -1
                })

        ws.delete_rows(index)
        wb.save(EXCEL_FILE)

        return jsonify({
            "success": True,
            "message": "Entry successfully deleted.",
            "index": index - 2
        })
    

    return jsonify({
        "success": False,
        "message": "Not a valid action.",
        "index": -1
    })
    

def binary_search(worksheet, requestUrl):
    left = 2
    right = find_index_of_last_row(worksheet)

    if right < 2:
        return -1

    while left <= right:
        mid = (left + right) // 2
        comparisonUrl = worksheet.cell(row=mid, column=1).value
        if comparisonUrl == requestUrl:
            return mid
        elif comparisonUrl < requestUrl:
            left = mid + 1
        else:
            right = mid - 1
    return -1


def binary_insert(worksheet, requestUrl):
    left = 2
    right = find_index_of_last_row(worksheet)

    if right < 2:
        return True, 2

    while left <= right:
        mid = (left + right) // 2
        comparisonUrl = worksheet.cell(row=mid, column=1).value
        if comparisonUrl == requestUrl:
            return False, mid
        elif comparisonUrl < requestUrl:
            left = mid + 1
        else:
            right = mid - 1
    return True, left


def find_index_of_last_row(worksheet):
    for row_idx in range(worksheet.max_row, 0, -1):
        row_is_empty = all(cell.value is None for cell in worksheet[row_idx])
        if not row_is_empty:
            # Found the real last non-empty row
            return row_idx
    # Sheet is empty
    return 0


if __name__ == '__main__':
    app.run(host='127.0.0.1', port=5001, debug=True)
